"""
Excel export module for property records.

This module generates formatted Excel workbooks with multiple sheets
as specified in FR-7.
"""

from typing import Dict, List, Any, Optional, TYPE_CHECKING
from pathlib import Path
from datetime import datetime

if TYPE_CHECKING:
    from openpyxl import Workbook

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore

from config.settings import EXCEL_CONFIG, OUTPUT_DIR
from src.utils import get_logger


class PropertyExporter:
    """Exporter for creating formatted Excel workbooks."""

    def __init__(self):
        """Initialize the exporter."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")

        self.logger = get_logger(__name__)
        self.config = EXCEL_CONFIG

    def export_to_excel(
        self,
        all_records: List[Dict[str, Any]],
        wake_records: List[Dict[str, Any]],
        orange_records: List[Dict[str, Any]],
        duplicates: List[Dict[str, Any]],
        statistics: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Path:
        """
        Export property records to a formatted Excel workbook.

        Creates a multi-sheet workbook with:
        - All Properties
        - Wake County
        - Orange County
        - Duplicates
        - Statistics

        Args:
            all_records: All property records (deduplicated)
            wake_records: Wake County records only
            orange_records: Orange County records only
            duplicates: Duplicate records
            statistics: Pipeline statistics
            filename: Optional output filename

        Returns:
            Path to created Excel file
        """
        self.logger.info("Creating Excel workbook...")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_data_sheet(wb, "All Properties", all_records)
        self._create_data_sheet(wb, "Wake County", wake_records)
        self._create_data_sheet(wb, "Orange County", orange_records)
        self._create_data_sheet(wb, "Duplicates", duplicates)
        self._create_statistics_sheet(wb, "Statistics", statistics)

        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.config['file_prefix']}_{timestamp}.xlsx"

        # Save workbook
        output_path = OUTPUT_DIR / filename
        wb.save(output_path)

        self.logger.info(f"Excel workbook created: {output_path}")
        return output_path

    def _create_data_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        records: List[Dict[str, Any]]
    ):
        """
        Create a data sheet with property records.

        Args:
            wb: Workbook object
            sheet_name: Name of the sheet
            records: List of property records
        """
        # Create sheet
        ws = wb.create_sheet(sheet_name)

        if not records:
            ws.append(["No records"])
            return

        # Get column headers from first record
        headers = self._get_column_headers(records[0])

        # Write headers
        ws.append(headers)

        # Format headers
        self._format_header_row(ws, len(headers))

        # Write data rows
        for record in records:
            row_data = [record.get(field, "") for field in headers]
            ws.append(row_data)

        # Format data
        self._format_data_sheet(ws, len(headers))

        # Add conditional formatting for quality scores
        if "quality_score" in headers:
            self._add_quality_score_formatting(ws, headers.index("quality_score") + 1, len(records))

        self.logger.debug(f"Created sheet '{sheet_name}' with {len(records)} records")

    def _get_column_headers(self, record: Dict[str, Any]) -> List[str]:
        """
        Get ordered column headers from a record.

        Args:
            record: Property record

        Returns:
            List of column headers
        """
        # Priority order for columns
        priority_fields = [
            "owner_name",
            "parcel_id",
            "property_address",
            "city",
            "state",
            "zip_code",
            "county",
            "mailing_address",
            "assessed_value",
            "sale_date",
            "sale_price",
            "quality_score",
            "quality_level",
            "completeness_percent",
            "source",
            "source_url",
            "extracted_at",
        ]

        # Get all fields
        all_fields = list(record.keys())

        # Order: priority fields first, then remaining fields alphabetically
        ordered_fields = []

        # Add priority fields that exist
        for field in priority_fields:
            if field in all_fields:
                ordered_fields.append(field)

        # Add remaining fields
        remaining_fields = sorted([f for f in all_fields if f not in ordered_fields])
        ordered_fields.extend(remaining_fields)

        return ordered_fields

    def _format_header_row(self, ws, num_columns: int):
        """
        Format header row.

        Args:
            ws: Worksheet object
            num_columns: Number of columns
        """
        formatting = self.config["formatting"]

        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)

            # Bold font
            if formatting["header_bold"]:
                cell.font = Font(bold=True)

            # Background color
            if formatting["header_background"]:
                cell.fill = PatternFill(
                    start_color=formatting["header_background"],
                    end_color=formatting["header_background"],
                    fill_type="solid"
                )

            # Alignment
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _format_data_sheet(self, ws, num_columns: int):
        """
        Format data sheet.

        Args:
            ws: Worksheet object
            num_columns: Number of columns
        """
        formatting = self.config["formatting"]

        # Set column widths
        for col in range(1, num_columns + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = formatting["column_width"]

        # Freeze panes
        if formatting["freeze_panes"]:
            ws.freeze_panes = formatting["freeze_panes"]

        # Auto filter
        if formatting["auto_filter"] and ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(num_columns)}{ws.max_row}"

    def _add_quality_score_formatting(
        self,
        ws,
        quality_column: int,
        num_records: int
    ):
        """
        Add conditional formatting for quality scores.

        Args:
            ws: Worksheet object
            quality_column: Column index for quality_score
            num_records: Number of data rows
        """
        if num_records == 0:
            return

        cond_format = self.config["conditional_formatting"]["quality_score"]

        # Apply color fills based on quality score
        for row in range(2, num_records + 2):  # Skip header
            cell = ws.cell(row=row, column=quality_column)

            try:
                score = float(cell.value) if cell.value else 0
            except (ValueError, TypeError):
                continue

            # Determine color based on score
            if score >= cond_format["high"]["min"]:
                color = cond_format["high"]["color"]
            elif score >= cond_format["medium"]["min"]:
                color = cond_format["medium"]["color"]
            else:
                color = cond_format["low"]["color"]

            # Apply fill
            cell.fill = PatternFill(
                start_color=color,
                end_color=color,
                fill_type="solid"
            )

    def _create_statistics_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        statistics: Dict[str, Any]
    ):
        """
        Create statistics summary sheet.

        Args:
            wb: Workbook object
            sheet_name: Name of the sheet
            statistics: Pipeline statistics dictionary
        """
        ws = wb.create_sheet(sheet_name)

        # Title
        ws.append(["Property Data Extraction Pipeline - Statistics"])
        ws.append([])

        # Format title
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.merge_cells("A1:B1")

        # Add statistics
        for section, values in statistics.items():
            # Section header
            ws.append([section.replace("_", " ").title()])
            ws.cell(ws.max_row, 1).font = Font(bold=True)

            # Section values
            if isinstance(values, dict):
                for key, value in values.items():
                    ws.append([f"  {key.replace('_', ' ').title()}", value])
            else:
                ws.append(["  Value", values])

            ws.append([])

        # Add generation timestamp
        ws.append(["Report Generated"])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.append(["  Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        # Set column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

        self.logger.debug(f"Created statistics sheet '{sheet_name}'")

    def export_to_csv(
        self,
        records: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> Path:
        """
        Export property records to CSV format.

        Args:
            records: List of property records
            filename: Optional output filename

        Returns:
            Path to created CSV file
        """
        import csv

        self.logger.info("Creating CSV export...")

        if not records:
            self.logger.warning("No records to export")
            return None

        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.config['file_prefix']}_{timestamp}.csv"

        output_path = OUTPUT_DIR / filename

        # Get headers
        headers = self._get_column_headers(records[0])

        # Write CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(records)

        self.logger.info(f"CSV file created: {output_path}")
        return output_path

    def export_to_json(
        self,
        records: List[Dict[str, Any]],
        filename: Optional[str] = None
    ) -> Path:
        """
        Export property records to JSON format.

        Args:
            records: List of property records
            filename: Optional output filename

        Returns:
            Path to created JSON file
        """
        import json

        self.logger.info("Creating JSON export...")

        # Generate filename if not provided
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.config['file_prefix']}_{timestamp}.json"

        output_path = OUTPUT_DIR / filename

        # Write JSON
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(records, f, indent=2, default=str)

        self.logger.info(f"JSON file created: {output_path}")
        return output_path
